from numpy.matlib import empty
from psycopg2 import Error
from sqlalchemy import create_engine, select, func, distinct
from sqlalchemy.orm import sessionmaker, joinedload, selectinload, join, DeclarativeBase
from sqlalchemy.util.queue import Empty

# from database.models.exercise import Сompleted_exercise, Exercise_structure
from database.services.education import education_service_db
from schemas.test import ResScale, ReqBorder, ReqScale
from typing import List
from sqlalchemy.exc import NoResultFound

from database.inquiries import inquiries
from database.models.client import *
from database.models.diary import *
from database.models.education import *
from database.models.experiment import *
from database.models.inquiry import *
from database.models.mood_tracker import *
from database.models.post import *
from database.models.problem import *
from database.models.test import *
from database.models.users import *
from database.models.review import *

from database.database import engine, session_factory
from database.calculator import calculator_service
from database.enum import DiaryType
from fastapi.responses import FileResponse
import uuid
from openpyxl import load_workbook
import pandas as pd


class UserStatisticsServicedb:
    def general_test_results(self, psychologist_id: uuid.UUID):
        with session_factory() as session:
            try:

                df = pd.DataFrame()
                excel_file = 'users.xlsx'
                df.to_excel(excel_file, index=False)

                fn = "users.xlsx"
                wb = load_workbook(fn)

                wb.create_sheet("выгорание")
                ws = wb["выгорание"]
                ws.append(["Email", "Эмоциональное истощение", "Деперсонализация", "Редукция проф. достижений", "Дата"])

                wb.create_sheet("DASS-21")
                ws = wb["DASS-21"]
                ws.append(["Email", "Тревога", "Депрессия", "Стресс", "Дата"])

                wb.create_sheet("тревоги")
                ws = wb["тревоги"]
                ws.append(["Email", "Шкала ситуативной тревожности", "Шкала личностной тревожности", "Дата"])

                wb.create_sheet("копинг")
                ws = wb["копинг"]
                ws.append(["Email", "Разрешение проблем", "Поиск социальной поддержки", "Избегание проблем", "Дата"])

                wb.create_sheet("CMQ")
                ws = wb["CMQ"]
                ws.append(["Email", "Шкала когнитивных ошибок", "Персонализация", "Чтение мыслей", "Упрямство",
                           "Морализация", "Катастрофизация", "Выученная беспомощность", "Максимализм",
                           "Преувеличение опасности", "Гипернормативность", "Дата"])

                wb.create_sheet("апатии")
                ws = wb["апатии"]
                ws.append(["Email", "Шкала профессиональной апатии", "Апатичные мысли", "Апатичные действия", "Дата"])

                wb.create_sheet("Бека")
                ws = wb["Бека"]
                ws.append(["Email", "Шкала депрессии", "Когнитивно-аффективная субшкала",
                           "Субшкала соматических проявлений депрессии", "Дата"])

                wb.create_sheet("Самооценка")
                ws = wb["Самооценка"]
                ws.append(["Email", "Шкала воспринимаемого стресса", "Фактор дистресса", "Фактор совладания", "Дата"])

                query = (
                    session.query(Clients)
                    .filter(Clients.psychologist_id == psychologist_id)
                    .join(Users, Clients.client_id == Users.id)
                    .join(Test_result, Test_result.user_id == Users.id)
                    .join(Test, Test_result.test_id == Test.id)
                    .options(
                        joinedload(Users.test_result)
                        .joinedload(Test_result.scale_result)
                        .joinedload(Scale_result.scale_id)
                    )
                )
                res = session.execute(query)
                clients = res.unique().scalars().all()
                users = session.query(Users).all()
                tests = session.query(Test).all()
                test_titles = {
                    'Профессиональное выгорание': "выгорание",
                    'DASS - 21': 'DASS-21',
                    'Шкала тревоги Спилбергера-Ханина, STAI': "тревоги",
                    'Индикатор копинг-стратегий': "копинг",
                    'Опросник когнтитвных ошибок CMQ': "CMQ",
                    'Шкала профессиональной апатии': "апатии",
                    'Шкала депрессии Бека': "Бека",
                    'Самооценка стрессоустойчивости Коухена-Виллиансона': "Самооценка"
                }

                user_map = {user.id: user for user in users}
                test_map = {test.id: test for test in tests}

                for client in clients:
                    user = user_map.get(client.client_id)
                    if not user:
                        continue

                    for test_result in user.test_result:
                        test = test_map.get(test_result.test_id)
                        if not test:
                            continue

                        test_title = test.title
                        ws = wb[test_titles[test_title]]

                        scores = [scale_result.score for scale_result in test_result.scale_result]

                        ws.append([user.email] + scores + [test_result.date.strftime("%d.%m.%Y")])

                wb.save(fn)
                wb.close()

                response = FileResponse(fn,
                                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                        filename=fn)

                return response

            except Exception as error:
                print(error)

    def user_activity_statistics(self):
        with session_factory() as session:
            try:
                completed_exercise = session.query(Сompleted_exercise).all()
                users = session.query(Users).all()

                df = pd.DataFrame()
                excel_file = 'userss.xlsx'
                df.to_excel(excel_file, index=False)
                fn = "userss.xlsx"
                wb = load_workbook(fn)
                wb.create_sheet("excel score")
                ws = wb["excel score"]
                ws.append(["Email", "Основы", "Основы КПТ", "Выгорание", "Дыхательные техники", "Техники релаксации",
                           "Копинг стратегии", "КПТ-дневник", "Трекер настроения", "КПТ-дневник", "Заметки", "Дата"])

                for user in users:
                    score_list = []
                    date_list = []
                    score_list_excercise = [0, 0, 0]
                    education_list = education_service_db.get_all_education_theme_db(str(user.id))
                    for education in education_list:
                        score_list.append(education['score'])
                    for exercise in completed_exercise:
                        if str(user.id) == str(exercise.user_id):
                            date_list.append(exercise.date)
                    date_set = set(date_list)

                    print("score_list1")
                    if not date_list:
                        ws.append([user.email] + score_list + score_list_excercise)
                    for date in date_set:
                        excel_date = ""
                        for exercise in completed_exercise:
                            if date == exercise.date and str(user.id) == str(exercise.user_id):
                                excel_date = date.strftime("%d.%m.%Y")
                                exercise_structure = session.get(Exercise_structure, exercise.exercise_structure_id)
                                if exercise_structure.title == "Трекер настроения":
                                    score_list_excercise[0] += 1
                                elif exercise_structure.title == "КПТ-дневник":
                                    score_list_excercise[1] += 1
                                elif exercise_structure.title == "Заметки":
                                    score_list_excercise[2] += 1
                        print(score_list)
                        print("score_list")
                        ws.append([user.email] + score_list + score_list_excercise + [excel_date])

                wb.save(fn)
                wb.close()

                response = FileResponse(fn,
                                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                        filename=fn)

                return response

            except Exception as error:
                print(error)


user_statistics_service_db: UserStatisticsServicedb = UserStatisticsServicedb()
